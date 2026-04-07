#!/usr/bin/env python3
"""
Генерирует продающие названия товаров для категории "Тарифные планы" в 44.xlsx.
Паттерн ПиплБот + правила Яндекс Маркета.

Формулы названий:
  eSIM:    "ESIM {страна} {трафик} на {срок}, мобильный интернет, виртуальная сим-карта"
  Voucher: "Ваучер {оператор} {номинал} {валюта}, {страна}, пополнение мобильной связи"
  Data:    "Пакет мобильного интернета {оператор} {трафик} на {срок}, {страна}"

Цель: 50-120 символов, уникальные, на русском, без дублирования.
"""

import openpyxl
import re
import os

# === МАППИНГ СТРАН (из бренда → русское название) ===
COUNTRY_MAP = {
    'SA': 'Саудовская Аравия', 'AE': 'ОАЭ', 'UK': 'Великобритания',
    'FR': 'Франция', 'DE': 'Германия', 'NL': 'Нидерланды',
    'OM': 'Оман', 'KW': 'Кувейт', 'JO': 'Иордания',
}

COUNTRY_FROM_BRAND = {
    'Afghanistan': 'Афганистан', 'Algeria': 'Алжир', 'Angola': 'Ангола',
    'Austria': 'Австрия', 'Bahrain': 'Бахрейн', 'Bangladesh': 'Бангладеш',
    'Belgium': 'Бельгия', 'China': 'Китай', 'Cyprus': 'Кипр',
    'Egypt': 'Египет', 'France': 'Франция', 'Georgia': 'Грузия',
    'Germany': 'Германия', 'Greece': 'Греция', 'Guinea': 'Гвинея',
    'Indonesia': 'Индонезия', 'Italy': 'Италия', 'Japan': 'Япония',
    'Montenegro': 'Черногория', 'Portugal': 'Португалия', 'Qatar': 'Катар',
    'Russia': 'Россия', 'Serbia': 'Сербия', 'Singapore': 'Сингапур',
    'Spain': 'Испания', 'Thailand': 'Таиланд', 'Tunisia': 'Тунис',
    'Turkey': 'Турция', 'Uzbekistan': 'Узбекистан', 'Vietnam': 'Вьетнам',
}

MULTI_WORD_COUNTRIES = {
    'Hong Kong': 'Гонконг', 'Saudi Arabia': 'Саудовская Аравия',
    'United Arab Emirates': 'ОАЭ', 'United Kingdom': 'Великобритания',
    'United States': 'США',
}

# Страна из описания (Страна: XX)
DESC_COUNTRY_CODES = {
    'AF': 'Афганистан', 'AE': 'ОАЭ', 'DZ': 'Алжир', 'AO': 'Ангола',
    'AT': 'Австрия', 'BH': 'Бахрейн', 'BD': 'Бангладеш', 'BE': 'Бельгия',
    'CN': 'Китай', 'CY': 'Кипр', 'EG': 'Египет', 'FR': 'Франция',
    'GE': 'Грузия', 'DE': 'Германия', 'GR': 'Греция', 'GN': 'Гвинея',
    'HK': 'Гонконг', 'ID': 'Индонезия', 'IT': 'Италия', 'JP': 'Япония',
    'JO': 'Иордания', 'KW': 'Кувейт', 'ME': 'Черногория', 'OM': 'Оман',
    'PT': 'Португалия', 'QA': 'Катар', 'RU': 'Россия', 'SA': 'Саудовская Аравия',
    'RS': 'Сербия', 'SG': 'Сингапур', 'ES': 'Испания', 'TH': 'Таиланд',
    'TN': 'Тунис', 'TR': 'Турция', 'GB': 'Великобритания', 'US': 'США',
    'UZ': 'Узбекистан', 'VN': 'Вьетнам',
}

# Валюта → полное название для ваучеров
CURRENCY_NAMES = {
    'SAR': 'SAR', 'AED': 'AED', 'KWD': 'KWD', 'OMR': 'OMR',
    'BHD': 'BHD', 'GBP': 'GBP', 'EUR': 'EUR', 'USD': 'USD',
    'AFN': 'AFN', 'BDT': 'BDT', 'DZD': 'DZD', 'AOA': 'AOA',
    'JOD': 'JOD', 'EGP': 'EGP', 'GNF': 'GNF', 'THB': 'THB',
    '$': 'USD',
}

# Оператор из бренда (убираем суффикс | XX и слова Voucher/eVoucher)
def extract_operator(brand):
    """Извлечь чистое имя оператора из бренда."""
    # Убираем "| SA", "| AE" и т.д.
    op = re.sub(r'\s*\|\s*\w+$', '', brand).strip()
    # Убираем "Voucher", "eVoucher" из конца
    op = re.sub(r'\s*(Voucher|eVoucher|Credit)$', '', op, flags=re.IGNORECASE).strip()
    # Убираем название страны из конца бренда
    for eng in list(COUNTRY_FROM_BRAND.keys()) + list(MULTI_WORD_COUNTRIES.keys()):
        if op.endswith(eng):
            op = op[:-(len(eng))].strip()
            break
    return op or brand


def extract_country(brand, name, desc):
    """Извлечь страну."""
    # 1. Суффикс бренда
    m = re.search(r'\|\s*(\w+)$', brand)
    if m and m.group(1) in COUNTRY_MAP:
        return COUNTRY_MAP[m.group(1)]

    # 2. Двусловная страна в бренде
    for eng, rus in MULTI_WORD_COUNTRIES.items():
        if eng in brand:
            return rus

    # 3. Однословная страна в бренде
    for eng, rus in COUNTRY_FROM_BRAND.items():
        if brand.endswith(eng) or f' {eng}' in brand:
            return rus

    # 4. Из описания
    m = re.search(r'Страна:\s*(\w+)', desc)
    if m and m.group(1) in DESC_COUNTRY_CODES:
        return DESC_COUNTRY_CODES[m.group(1)]

    # 5. Airalo
    if 'airalo' in brand.lower():
        return 'весь мир'

    return None


def extract_denomination(name):
    """Извлечь номинал: (валюта, сумма)."""
    # Стандартный: "SAR 34.50", "AED 25", "$5"
    currencies = r'SAR|AED|KWD|OMR|BHD|GBP|EUR|USD|AFN|BDT|DZD|AOA|JOD|EGP|GNF|THB'
    m = re.search(rf'({currencies}|\$)\s*([\d.,]+)', name)
    if m:
        cur = m.group(1)
        if cur == '$':
            cur = 'USD'
        amt = m.group(2).replace(',', '.')
        # Убираем trailing dot
        amt = amt.rstrip('.')
        return cur, amt

    # Реверс: "30 EUR"
    m = re.search(rf'([\d.,]+)\s*({currencies})', name)
    if m:
        amt = m.group(1).replace(',', '.').rstrip('.')
        return m.group(2), amt

    return None, None


def extract_traffic(name, desc):
    """Извлечь трафик."""
    combined = name + ' ' + desc
    # "10GB", "1.5GB", "2ГБ", "500MB"
    m = re.search(r'([\d.]+)\s*GB', combined, re.IGNORECASE)
    if m:
        val = m.group(1)
        # Если дробное — оставляем как есть, иначе убираем .0
        if '.' in val:
            return f'{val} ГБ'
        return f'{val} ГБ'
    m = re.search(r'(\d+)\s*MB', combined, re.IGNORECASE)
    if m:
        return f'{m.group(1)} МБ'
    m = re.search(r'([\d.]+)\s*ГБ', combined)
    if m:
        return f'{m.group(1)} ГБ'
    # "Unlimited"
    if re.search(r'unlimited|безлимит', combined, re.IGNORECASE):
        return 'безлимитный'
    return None


def extract_days(name, desc):
    """Извлечь срок в днях."""
    combined = name + ' ' + desc
    m = re.search(r'(\d+)\s*(?:Days?|дн)', combined, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def format_days_text(days):
    """Форматировать дни."""
    if days is None:
        return None
    if days == 1:
        return '1 день'
    elif days in (2, 3, 4):
        return f'{days} дня'
    else:
        return f'{days} дней'


# Ключевые слова для ротации (добавляют уникальность)
KEYWORDS_ESIM = [
    'цифровой код',
    'онлайн активация',
    'электронная сим-карта',
    'мгновенная доставка',
    'без физической карты',
]

KEYWORDS_VOUCHER = [
    'цифровой код',
    'электронный ваучер',
    'мгновенная доставка',
    'онлайн пополнение',
    'электронное пополнение',
]


def generate_esim_name(country, traffic, days_text, sku_idx):
    """Генерировать название для eSIM."""
    # Формула: ESIM {страна} {трафик} на {срок}, мобильный интернет, {ключевое}
    parts = ['ESIM']

    if country:
        parts.append(country)

    if traffic:
        if days_text:
            parts.append(f'{traffic} на {days_text}')
        else:
            parts.append(traffic)
    elif days_text:
        parts.append(f'на {days_text}')

    base = ' '.join(parts)

    # Добавляем контекст
    suffix = 'мобильный интернет, виртуальная сим-карта'
    name = f'{base}, {suffix}'

    # Если слишком длинное (>120), сокращаем
    if len(name) > 120:
        name = f'{base}, мобильный интернет'

    # Добавляем ключевое слово для уникальности если название повторяется
    kw = KEYWORDS_ESIM[sku_idx % len(KEYWORDS_ESIM)]

    # Проверяем длину с ключевым словом
    name_with_kw = f'{base}, {suffix}, {kw}'
    if len(name_with_kw) <= 130:
        name = name_with_kw

    return name


def generate_voucher_name(operator, country, currency, amount, sku_idx):
    """Генерировать название для ваучера."""
    # Формула: Ваучер {оператор} {номинал} {валюта}, {страна}, пополнение мобильной связи
    parts = ['Ваучер']

    if operator:
        parts.append(operator)

    if amount and currency:
        parts.append(f'{amount} {currency}')

    base = ' '.join(parts)

    if country and country != 'весь мир':
        name = f'{base}, {country}, пополнение мобильной связи'
    elif country == 'весь мир':
        name = f'{base}, международный роуминг, пополнение связи'
    else:
        name = f'{base}, пополнение мобильной связи'

    # Если слишком короткое (<50), добавляем ключевое слово
    if len(name) < 50:
        kw = KEYWORDS_VOUCHER[sku_idx % len(KEYWORDS_VOUCHER)]
        name = f'{name}, {kw}'

    # Если слишком длинное (>130), сокращаем
    if len(name) > 130:
        if country:
            name = f'{base}, {country}'
        else:
            name = base

    return name


def generate_data_pack_name(operator, country, traffic, days_text, sku_idx):
    """Генерировать название для дата-пакета (AIS Thailand, Banglalink Data и т.д.)."""
    parts = ['Пакет мобильного интернета']

    if operator:
        parts.append(operator)

    if traffic:
        if days_text:
            parts.append(f'{traffic} на {days_text}')
        else:
            parts.append(traffic)
    elif days_text:
        parts.append(f'на {days_text}')

    base = ' '.join(parts)

    if country:
        name = f'{base}, {country}'
    else:
        name = base

    # Добавляем ключевое слово для длины
    if len(name) < 50:
        kw = KEYWORDS_VOUCHER[sku_idx % len(KEYWORDS_VOUCHER)]
        name = f'{name}, {kw}'

    return name


def is_data_pack(brand, name):
    """Определить, является ли товар дата-пакетом (не eSIM и не простой ваучер)."""
    return 'Data Pack' in name or ('Data Credit' in name and 'GB' in name.upper())


def main():
    src = os.path.join(os.path.dirname(__file__), '..', '..', 'new table', '44_filled.xlsx')
    src = os.path.abspath(src)
    dst = src  # перезаписываем тот же файл

    print(f'Читаю: {src}')
    wb = openpyxl.load_workbook(src)
    ws = wb['Данные о товарах']

    stats = {'total': 0, 'esim': 0, 'voucher': 0, 'data_pack': 0}
    names_generated = []
    name_counts = {}

    for row_idx in range(8, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=1).value
        if not sku or not str(sku).startswith('GFT'):
            continue

        old_name = str(ws.cell(row=row_idx, column=7).value or '')
        brand = str(ws.cell(row=row_idx, column=11).value or '')
        desc = str(ws.cell(row=row_idx, column=10).value or '')

        stats['total'] += 1
        sku_idx = stats['total']

        country = extract_country(brand, old_name, desc)
        is_esim = 'esim' in brand.lower()

        if is_esim:
            # eSIM товар
            traffic = extract_traffic(old_name, desc)
            days = extract_days(old_name, desc)
            days_text = format_days_text(days)

            new_name = generate_esim_name(country, traffic, days_text, sku_idx)
            stats['esim'] += 1

        elif is_data_pack(brand, old_name):
            # Дата-пакет оператора
            operator = extract_operator(brand)
            traffic = extract_traffic(old_name, desc)
            days = extract_days(old_name, desc)
            days_text = format_days_text(days)

            new_name = generate_data_pack_name(operator, country, traffic, days_text, sku_idx)
            stats['data_pack'] += 1

        else:
            # Ваучер пополнения
            operator = extract_operator(brand)
            currency, amount = extract_denomination(old_name)
            # Если не нашли в названии, пробуем в описании
            if not currency:
                currency, amount = extract_denomination(desc)

            new_name = generate_voucher_name(operator, country, currency, amount, sku_idx)
            stats['voucher'] += 1

        # Track duplicates
        name_counts[new_name] = name_counts.get(new_name, 0) + 1

        # Write new name
        ws.cell(row=row_idx, column=7).value = new_name
        names_generated.append((str(sku), old_name[:60], new_name, len(new_name)))

    # Resolve duplicates by adding different keywords
    if any(c > 1 for c in name_counts.values()):
        dup_counter = {}
        for row_idx in range(8, ws.max_row + 1):
            sku = ws.cell(row=row_idx, column=1).value
            if not sku or not str(sku).startswith('GFT'):
                continue
            name = ws.cell(row=row_idx, column=7).value
            if name_counts.get(name, 0) > 1:
                idx = dup_counter.get(name, 0)
                dup_counter[name] = idx + 1
                # Add unique suffix based on index
                brand = str(ws.cell(row=row_idx, column=11).value or '')
                is_esim = 'esim' in brand.lower()
                kw_list = KEYWORDS_ESIM if is_esim else KEYWORDS_VOUCHER
                kw = kw_list[idx % len(kw_list)]
                if kw not in name:
                    new_name = f'{name}, {kw}'
                    if len(new_name) <= 140:
                        ws.cell(row=row_idx, column=7).value = new_name

    # Save
    print(f'Сохраняю: {dst}')
    wb.save(dst)
    wb.close()

    # Report
    print(f'\n=== Статистика ===')
    print(f'Всего: {stats["total"]}')
    print(f'  eSIM: {stats["esim"]}')
    print(f'  Ваучеры: {stats["voucher"]}')
    print(f'  Дата-пакеты: {stats["data_pack"]}')

    # Length stats
    lengths = [n[3] for n in names_generated]
    print(f'\nДлина названий:')
    print(f'  Мин: {min(lengths)}, Макс: {max(lengths)}, Средняя: {sum(lengths)//len(lengths)}')
    print(f'  < 50 символов: {sum(1 for l in lengths if l < 50)}')
    print(f'  50-120 символов: {sum(1 for l in lengths if 50 <= l <= 120)}')
    print(f'  > 120 символов: {sum(1 for l in lengths if l > 120)}')

    # Duplicates
    dups = sum(1 for c in name_counts.values() if c > 1)
    print(f'\nДубликаты названий: {dups}')

    # Samples
    print(f'\n=== Примеры (15 штук) ===')
    for sku, old, new, length in names_generated[:5]:
        print(f'  {sku}: ({length} сим)')
        print(f'    БЫЛО:  {old}')
        print(f'    СТАЛО: {new}')
    for sku, old, new, length in names_generated[60:65]:
        print(f'  {sku}: ({length} сим)')
        print(f'    БЫЛО:  {old}')
        print(f'    СТАЛО: {new}')
    for sku, old, new, length in names_generated[200:205]:
        print(f'  {sku}: ({length} сим)')
        print(f'    БЫЛО:  {old}')
        print(f'    СТАЛО: {new}')


if __name__ == '__main__':
    main()
