#!/usr/bin/env python3
"""
Генерирует HTML-описания для 44_filled.xlsx по структуре из 3.xlsx:
  1. <h2>Заголовок — что это</h2>
  2. <p>Вводный абзац — что даёт продукт</p>
  3. <p><b>Что вы получите / Преимущества:</b></p><ul>...</ul>
  4. <p><b>Обратите внимание:</b></p><ul>...</ul>
  5. <h2>Инструкция по активации</h2> + пошаговые <ol>

Инструкция eSIM — с ПиплБот (iPhone + Android + проверка).
Инструкция ваучеров — USSD / приложение оператора.
"""

import openpyxl
import re
import os

# === МАППИНГИ ===
COUNTRY_MAP = {
    'SA': 'Саудовская Аравия', 'AE': 'ОАЭ', 'UK': 'Великобритания',
    'FR': 'Франция', 'DE': 'Германия', 'NL': 'Нидерланды',
    'OM': 'Оман', 'KW': 'Кувейт', 'JO': 'Иордания',
}
COUNTRY_FROM_BRAND = {
    'Afghanistan': 'Афганистан', 'Algeria': 'Алжир', 'Angola': 'Ангола',
    'Austria': 'Австрия', 'Bahrain': 'Бахрейн', 'Bangladesh': 'Бангладеш',
    'Belgium': 'Бельгия', 'China': 'Китай', 'Cyprus': 'Кипр',
    'Egypt': 'Египет', 'France': 'Франция', 'Georgia': 'Грузия',
    'Germany': 'Германия', 'Greece': 'Греция', 'Guinea': 'Гвинея',
    'Indonesia': 'Индонезия', 'Italy': 'Италия', 'Japan': 'Япония',
    'Montenegro': 'Черногория', 'Portugal': 'Португалия', 'Qatar': 'Катар',
    'Russia': 'Россия', 'Serbia': 'Сербия', 'Singapore': 'Сингапур',
    'Spain': 'Испания', 'Thailand': 'Таиланд', 'Tunisia': 'Тунис',
    'Turkey': 'Турция', 'Uzbekistan': 'Узбекистан', 'Vietnam': 'Вьетнам',
}
MULTI_WORD_COUNTRIES = {
    'Hong Kong': 'Гонконг', 'Saudi Arabia': 'Саудовская Аравия',
    'United Arab Emirates': 'ОАЭ', 'United Kingdom': 'Великобритания',
    'United States': 'США',
}
DESC_COUNTRY_CODES = {
    'AF': 'Афганистан', 'AE': 'ОАЭ', 'DZ': 'Алжир', 'AO': 'Ангола',
    'AT': 'Австрия', 'BH': 'Бахрейн', 'BD': 'Бангладеш', 'BE': 'Бельгия',
    'CN': 'Китай', 'CY': 'Кипр', 'EG': 'Египет', 'FR': 'Франция',
    'GE': 'Грузия', 'DE': 'Германия', 'GR': 'Греция', 'GN': 'Гвинея',
    'HK': 'Гонконг', 'ID': 'Индонезия', 'IT': 'Италия', 'JP': 'Япония',
    'JO': 'Иордания', 'KW': 'Кувейт', 'ME': 'Черногория', 'OM': 'Оман',
    'PT': 'Португалия', 'QA': 'Катар', 'RU': 'Россия', 'SA': 'Саудовская Аравия',
    'RS': 'Сербия', 'SG': 'Сингапур', 'ES': 'Испания', 'TH': 'Таиланд',
    'TN': 'Тунис', 'TR': 'Турция', 'GB': 'Великобритания', 'US': 'США',
    'UZ': 'Узбекистан', 'VN': 'Вьетнам',
}
COUNTRY_PREPOSITIONAL = {
    'Афганистан': 'в Афганистане', 'Алжир': 'в Алжире', 'Ангола': 'в Анголе',
    'Австрия': 'в Австрии', 'Бахрейн': 'в Бахрейне', 'Бангладеш': 'в Бангладеш',
    'Бельгия': 'в Бельгии', 'Китай': 'в Китае', 'Кипр': 'на Кипре',
    'Египет': 'в Египте', 'Франция': 'во Франции', 'Грузия': 'в Грузии',
    'Германия': 'в Германии', 'Греция': 'в Греции', 'Гвинея': 'в Гвинее',
    'Гонконг': 'в Гонконге', 'Индонезия': 'в Индонезии', 'Италия': 'в Италии',
    'Япония': 'в Японии', 'Иордания': 'в Иордании', 'Кувейт': 'в Кувейте',
    'Черногория': 'в Черногории', 'Оман': 'в Омане', 'Португалия': 'в Португалии',
    'Катар': 'в Катаре', 'Россия': 'в России', 'Саудовская Аравия': 'в Саудовской Аравии',
    'Сербия': 'в Сербии', 'Сингапур': 'в Сингапуре', 'Испания': 'в Испании',
    'Таиланд': 'в Таиланде', 'Тунис': 'в Тунисе', 'Турция': 'в Турции',
    'Великобритания': 'в Великобритании', 'США': 'в США',
    'Узбекистан': 'в Узбекистане', 'Вьетнам': 'во Вьетнаме',
    'ОАЭ': 'в ОАЭ', 'Нидерланды': 'в Нидерландах',
    'весь мир': 'в любой стране мира',
}
COUNTRY_GENITIVE = {
    'Афганистан': 'Афганистана', 'Алжир': 'Алжира', 'Ангола': 'Анголы',
    'Австрия': 'Австрии', 'Бахрейн': 'Бахрейна', 'Бангладеш': 'Бангладеш',
    'Бельгия': 'Бельгии', 'Китай': 'Китая', 'Кипр': 'Кипра',
    'Египет': 'Египта', 'Франция': 'Франции', 'Грузия': 'Грузии',
    'Германия': 'Германии', 'Греция': 'Греции', 'Гвинея': 'Гвинеи',
    'Гонконг': 'Гонконга', 'Индонезия': 'Индонезии', 'Италия': 'Италии',
    'Япония': 'Японии', 'Иордания': 'Иордании', 'Кувейт': 'Кувейта',
    'Черногория': 'Черногории', 'Оман': 'Омана', 'Португалия': 'Португалии',
    'Катар': 'Катара', 'Россия': 'России', 'Саудовская Аравия': 'Саудовской Аравии',
    'Сербия': 'Сербии', 'Сингапур': 'Сингапура', 'Испания': 'Испании',
    'Таиланд': 'Таиланда', 'Тунис': 'Туниса', 'Турция': 'Турции',
    'Великобритания': 'Великобритании', 'США': 'США',
    'Узбекистан': 'Узбекистана', 'Вьетнам': 'Вьетнама',
    'ОАЭ': 'ОАЭ', 'Нидерланды': 'Нидерландов',
    'весь мир': 'любой страны',
}


def extract_country(brand, old_desc):
    m = re.search(r'\|\s*(\w+)$', brand)
    if m and m.group(1) in COUNTRY_MAP:
        return COUNTRY_MAP[m.group(1)]
    for eng, rus in MULTI_WORD_COUNTRIES.items():
        if eng in brand: return rus
    for eng, rus in COUNTRY_FROM_BRAND.items():
        if brand.endswith(eng) or f' {eng}' in brand: return rus
    m = re.search(r'Страна:\s*(\w+)', old_desc)
    if m and m.group(1) in DESC_COUNTRY_CODES:
        return DESC_COUNTRY_CODES[m.group(1)]
    if 'airalo' in brand.lower(): return 'весь мир'
    return None


def extract_operator(brand):
    op = re.sub(r'\s*\|\s*\w+$', '', brand).strip()
    op = re.sub(r'\s*(Voucher|eVoucher|Credit)$', '', op, flags=re.IGNORECASE).strip()
    for eng in list(COUNTRY_FROM_BRAND.keys()) + list(MULTI_WORD_COUNTRIES.keys()):
        if op.endswith(eng): op = op[:-(len(eng))].strip()
    return op or brand


def extract_denomination(text):
    currencies = r'SAR|AED|KWD|OMR|BHD|GBP|EUR|USD|AFN|BDT|DZD|AOA|JOD|EGP|GNF|THB'
    m = re.search(rf'({currencies}|\$)\s*([\d.,]+)', text)
    if m:
        cur = 'USD' if m.group(1) == '$' else m.group(1)
        return cur, m.group(2).replace(',', '.').rstrip('.')
    m = re.search(rf'([\d.,]+)\s*({currencies})', text)
    if m: return m.group(2), m.group(1).replace(',', '.').rstrip('.')
    return None, None


def extract_traffic(name):
    m = re.search(r'([\d.]+)\s*ГБ', name)
    if m: return f'{m.group(1)} ГБ'
    m = re.search(r'([\d.]+)\s*МБ', name)
    if m: return f'{m.group(1)} МБ'
    if 'безлимитный' in name.lower(): return 'безлимитный'
    return None


def extract_days_from_name(name):
    m = re.search(r'(\d+)\s*(день|дня|дней)', name)
    if m: return f'{m.group(1)} {m.group(2)}'
    return None


# ============================================================
# ВВОДНЫЕ ФРАЗЫ (ротация)
# ============================================================
ESIM_INTROS = [
    'Виртуальная SIM-карта (eSIM) для мобильного интернета {where}. Мгновенная доставка QR-кода на электронную почту после покупки. Подключайтесь к сети за несколько минут без замены физической SIM-карты.',
    'Электронная SIM-карта для подключения к мобильному интернету {where}. После оплаты вы мгновенно получите QR-код для активации. eSIM работает параллельно с обычной SIM-картой.',
    'Цифровая SIM-карта для доступа в интернет {where}. Активируется через QR-код, не требует физической установки. Удобна для путешествий и командировок.',
    'eSIM для мобильного интернета {where} — современная альтернатива обычной SIM-карте. Мгновенная доставка кода после покупки. Работает на устройствах с поддержкой eSIM.',
    'Виртуальная сим-карта для выхода в интернет {where}. Устанавливается дистанционно через QR-код. Не занимает слот физической SIM-карты, удобна как вторая линия связи.',
]

VOUCHER_INTROS = [
    'Электронный ваучер для пополнения баланса оператора {operator} {where}. Мгновенная доставка кода на электронную почту после покупки. Пополняйте баланс без визита в салон связи.',
    'Цифровой код для пополнения мобильного счёта {operator} {where}. После оплаты вы мгновенно получите код активации. Удобный способ оплатить услуги связи онлайн.',
    'Ваучер пополнения для абонентов {operator} {where}. Моментальная доставка электронного кода после покупки. Средства зачисляются на баланс сразу после активации.',
    'Код пополнения для сети {operator} {where}. Мгновенная доставка на электронную почту. Пополните баланс телефона из любой точки мира без банковской карты оператора.',
    'Электронный код пополнения для оператора {operator} {where}. После покупки вы получите код на электронную почту. Активируйте через телефон или приложение оператора.',
]

DATA_INTROS = [
    'Пакет мобильного интернета от оператора {operator} для подключения {where}. Мгновенная доставка кода активации на электронную почту после покупки.',
    'Интернет-пакет {operator} {where}. После оплаты вы мгновенно получите код для активации трафика. Подключение через приложение оператора или USSD-команду.',
    'Дата-пакет оператора {operator} для мобильного интернета {where}. Моментальная доставка электронного кода после покупки. Трафик доступен сразу после активации.',
]


# ============================================================
# ИНСТРУКЦИЯ eSIM (по образцу ПиплБот)
# ============================================================
ESIM_INSTRUCTION = """<h2>Инструкция по активации eSIM</h2>
<p><b>1. Проверьте поддержку eSIM на вашем устройстве</b></p>
<p>Прежде чем начать, убедитесь, что ваше устройство поддерживает eSIM. Зайдите в настройки устройства и найдите раздел «Мобильная сеть» или «SIM-карты». Если устройство поддерживает eSIM, вы увидите опцию для добавления eSIM.</p>
<p><b>2. Подготовка к активации</b></p>
<ul>
<li>Убедитесь, что устройство подключено к интернету через Wi-Fi или другую мобильную сеть</li>
<li>Откройте настройки вашего устройства</li>
</ul>
<p><b>3. Добавьте eSIM на iPhone (iOS):</b></p>
<ol>
<li>Перейдите в «Настройки» — «Мобильные данные» — «Добавить тариф»</li>
<li>Выберите «Использовать QR-код»</li>
<li>Отсканируйте QR-код, отправленный на электронную почту, камерой iPhone</li>
<li>Следуйте инструкциям на экране для завершения активации</li>
<li>Убедитесь, что новая eSIM выбрана для мобильных данных</li>
</ol>
<p><b>4. Добавьте eSIM на Android:</b></p>
<ol>
<li>Перейдите в «Настройки» — «Подключения» — «SIM-карты» — «Добавить eSIM»</li>
<li>Выберите «Добавить eSIM с помощью QR-кода»</li>
<li>Отсканируйте QR-код из полученного письма камерой телефона</li>
<li>Следуйте инструкциям на экране для завершения активации</li>
<li>Убедитесь, что новая eSIM выбрана для мобильных данных</li>
</ol>
<p><b>5. Проверка активации</b></p>
<p>После завершения настройки проверьте подключение к интернету. Убедитесь, что eSIM выбрана в настройках в качестве основного оператора для мобильных данных.</p>
<p><b>Если устройство не подключается к интернету:</b></p>
<ul>
<li>Включите роуминг данных: «Настройки» — «Мобильные данные» — «Роуминг данных»</li>
<li>Убедитесь, что eSIM выбрана как основной источник для интернета</li>
<li>Попробуйте выбрать оператора вручную в настройках мобильных сетей</li>
<li>Включите и выключите авиарежим для восстановления связи с сетью</li>
</ul>"""


# ============================================================
# USSD-КОДЫ ОПЕРАТОРОВ (реальные, из официальных источников)
# ============================================================
OPERATOR_USSD = {
    'Mobily': {'ussd': '*1400*код_ваучера#', 'sms': 'код_ваучера на номер 1100', 'hotline': '1100', 'app': 'Mobily'},
    'Zain': {'ussd': '*141*код_ваучера#', 'sms': 'код_ваучера на номер 700212', 'hotline': '', 'app': 'Zain KSA'},
    'stc': {'ussd': '*155*код_ваучера#', 'sms': 'код_ваучера (пробел) 155 на номер 900', 'hotline': '1500', 'app': 'MySTC'},
    'Lebara': {'ussd': '*131*код_ваучера#', 'sms': '', 'hotline': '', 'app': 'MyLebara'},
    'Virgin Mobile': {'ussd': '*101*код_ваучера#', 'sms': '', 'hotline': '1789', 'app': 'Virgin Mobile'},
    'FRiENDi': {'ussd_sa': '*101*код_ваучера#', 'ussd_om': '*102*код_ваучера#', 'app': 'FRiENDi mobile'},
    'Ooredoo': {'ussd_om': '*999*код_ваучера#', 'app': 'Ooredoo'},
    'etisalat': {'ussd': '*120*код_ваучера#', 'app': 'e& (Etisalat)'},
    'Etisalat': {'ussd': '*120*код_ваучера#', 'app': 'e& (Etisalat)'},
    'du': {'ussd': '', 'app': 'du'},
    'Salam': {'ussd': '*101*код_ваучера#', 'app': 'Salam Mobile'},
    'Syma': {'ussd': '', 'app': 'Syma Mobile'},
    'Airalo': {'ussd': '', 'app': 'Airalo'},
}


def voucher_instruction(operator, country=None):
    info = OPERATOR_USSD.get(operator, {})
    ussd = info.get('ussd', '')
    # FRiENDi has different codes per country
    if operator == 'FRiENDi':
        if country and 'Оман' in country:
            ussd = info.get('ussd_om', '*102*код_ваучера#')
        else:
            ussd = info.get('ussd_sa', '*101*код_ваучера#')
    if operator == 'Ooredoo':
        ussd = info.get('ussd_om', '')
    sms = info.get('sms', '')
    hotline = info.get('hotline', '')
    app = info.get('app', operator)

    # Airalo — special case (not USSD, it's an app/website)
    if operator == 'Airalo':
        return f"""<h2>Инструкция по активации ваучера Airalo</h2>
<p><b>Способ 1 — Через приложение Airalo:</b></p>
<ol>
<li>Откройте приложение Airalo на телефоне и войдите в аккаунт</li>
<li>Перейдите в Профиль — Airmoney и Членство</li>
<li>Нажмите «Активировать ваучер» (Redeem Voucher)</li>
<li>Введите полученный код и нажмите «Активировать»</li>
<li>Средства будут зачислены на баланс Airmoney</li>
</ol>
<p><b>Способ 2 — Через сайт airalo.com:</b></p>
<ol>
<li>Откройте сайт airalo.com и войдите в аккаунт</li>
<li>Нажмите на имя профиля — выберите Airmoney и Членство</li>
<li>Нажмите «Активировать ваучер», введите код и подтвердите</li>
</ol>"""

    parts = []
    parts.append(f'<h2>Инструкция по активации кода {operator}</h2>')

    if ussd:
        parts.append(f'<p><b>Способ 1 — Через USSD-команду:</b></p>')
        parts.append('<ol>')
        parts.append('<li>Откройте приложение для набора номера на телефоне</li>')
        parts.append(f'<li>Наберите {ussd}, заменив «код_ваучера» на полученный код</li>')
        parts.append('<li>Нажмите кнопку вызова</li>')
        parts.append('<li>Баланс будет пополнен мгновенно, вы получите SMS-подтверждение</li>')
        parts.append('</ol>')

    if sms:
        parts.append(f'<p><b>Способ {2 if ussd else 1} — Через SMS:</b></p>')
        parts.append('<ol>')
        parts.append(f'<li>Отправьте SMS с текстом: {sms}</li>')
        parts.append('<li>Дождитесь SMS-подтверждения о зачислении средств</li>')
        parts.append('</ol>')

    app_num = 2 if (ussd and not sms) else (3 if ussd and sms else 1)
    parts.append(f'<p><b>Способ {app_num} — Через приложение {app}:</b></p>')
    parts.append('<ol>')
    parts.append(f'<li>Откройте приложение {app} на телефоне и войдите в аккаунт</li>')
    parts.append('<li>Перейдите в раздел «Пополнить баланс» или «Активировать код»</li>')
    parts.append('<li>Введите полученный код и подтвердите</li>')
    parts.append('<li>Средства зачислятся на баланс автоматически</li>')
    parts.append('</ol>')

    return '\n'.join(parts)


def data_instruction(operator, country=None):
    """Инструкция для дата-пакета — используем ту же логику с USSD."""
    info = OPERATOR_USSD.get(operator, {})
    ussd = info.get('ussd', '')
    if operator == 'FRiENDi':
        ussd = info.get('ussd_sa', '*101*код_ваучера#')
    if operator == 'Ooredoo':
        ussd = info.get('ussd_om', '')
    app = info.get('app', operator)

    parts = []
    parts.append(f'<h2>Инструкция по активации интернет-пакета {operator}</h2>')

    if ussd:
        parts.append('<p><b>Способ 1 — Через USSD-команду:</b></p>')
        parts.append('<ol>')
        parts.append('<li>Откройте приложение для набора номера на телефоне</li>')
        parts.append(f'<li>Наберите {ussd}, заменив «код_ваучера» на полученный код</li>')
        parts.append('<li>Нажмите кнопку вызова</li>')
        parts.append('<li>Интернет-трафик будет доступен сразу после активации</li>')
        parts.append('</ol>')

    parts.append(f'<p><b>Способ {2 if ussd else 1} — Через приложение {app}:</b></p>')
    parts.append('<ol>')
    parts.append(f'<li>Откройте приложение {app} на телефоне</li>')
    parts.append('<li>Перейдите в раздел «Пакеты» или «Активировать код»</li>')
    parts.append('<li>Введите полученный код и подтвердите активацию</li>')
    parts.append('<li>Трафик будет доступен сразу после активации</li>')
    parts.append('</ol>')

    return '\n'.join(parts)


# ============================================================
# ГЕНЕРАЦИЯ ОПИСАНИЙ
# ============================================================

def gen_esim(country, traffic, days_text, idx):
    where = COUNTRY_PREPOSITIONAL.get(country, f'в регионе {country}') if country else ''
    country_gen = COUNTRY_GENITIVE.get(country, country) if country else ''

    # h2 заголовок
    h2_parts = ['ESIM для мобильного интернета']
    if traffic: h2_parts.append(traffic)
    if country: h2_parts.append(f'— {country}')
    h2 = ' '.join(h2_parts)

    intro = ESIM_INTROS[idx % len(ESIM_INTROS)].format(where=where)

    # Преимущества
    benefits = [
        'Без физической карты — активация через QR-код',
        'Мгновенная доставка кода на электронную почту',
        'Работает параллельно с обычной SIM-картой',
    ]
    if traffic and traffic != 'безлимитный':
        benefits.append(f'Объём трафика: {traffic}')
    elif traffic == 'безлимитный':
        benefits.append('Безлимитный мобильный интернет')
    if days_text:
        benefits.append(f'Срок действия: {days_text} с момента активации')
    benefits.append('Совместим с iPhone XS и новее, Samsung Galaxy S20 и новее, Google Pixel 3 и новее')

    benefits_html = '\n'.join(f'<li>{b}</li>' for b in benefits)

    # Обратите внимание
    notes = []
    if country and country != 'весь мир':
        notes.append(f'eSIM предназначена для использования {where}')
    notes.append('Перед покупкой убедитесь, что ваше устройство поддерживает eSIM')
    notes.append('Каждый QR-код можно использовать только один раз')
    if days_text:
        notes.append(f'Срок действия тарифа — {days_text} с момента первого подключения к сети')
    notes_html = '\n'.join(f'<li>{n}</li>' for n in notes)

    return f"""<h2>{h2}</h2>
<p>{intro}</p>
<p><b>Преимущества виртуальной SIM-карты:</b></p>
<ul>
{benefits_html}
</ul>
<p><b>Обратите внимание:</b></p>
<ul>
{notes_html}
</ul>

{ESIM_INSTRUCTION}"""


def gen_voucher(operator, country, currency, amount, idx):
    where = COUNTRY_PREPOSITIONAL.get(country, f'в регионе {country}') if country else ''
    nominal = f'{amount} {currency}' if amount and currency else ''

    # h2
    h2 = f'Ваучер пополнения {operator}'
    if nominal: h2 += f' {nominal}'
    if country and country != 'весь мир': h2 += f' — {country}'

    intro = VOUCHER_INTROS[idx % len(VOUCHER_INTROS)].format(operator=operator, where=where)

    # Что вы получите
    items = []
    if nominal: items.append(f'Код пополнения на сумму {nominal}')
    items.append(f'Мгновенная доставка кода на электронную почту')
    items.append(f'Пополнение баланса оператора {operator} {where}')
    items.append('Код активации в электронном формате')
    items.append('Баланс зачисляется мгновенно после ввода кода')
    items_html = '\n'.join(f'<li>{i}</li>' for i in items)

    # Обратите внимание
    notes = []
    if country and country != 'весь мир':
        notes.append(f'Ваучер действует только для абонентов {operator} {where}')
    elif country == 'весь мир':
        notes.append('Ваучер действует для международного использования')
    notes.append('Код активации одноразовый и не подлежит возврату')
    notes.append('Срок действия кода не ограничен (если не указано иное)')
    notes_html = '\n'.join(f'<li>{n}</li>' for n in notes)

    return f"""<h2>{h2}</h2>
<p>{intro}</p>
<p><b>Что вы получите:</b></p>
<ul>
{items_html}
</ul>
<p><b>Обратите внимание:</b></p>
<ul>
{notes_html}
</ul>

{voucher_instruction(operator, country)}"""


def gen_data(operator, country, traffic, days_text, idx):
    where = COUNTRY_PREPOSITIONAL.get(country, f'в регионе {country}') if country else ''

    h2 = f'Пакет мобильного интернета {operator}'
    if traffic: h2 += f' {traffic}'
    if country: h2 += f' — {country}'

    intro = DATA_INTROS[idx % len(DATA_INTROS)].format(
        operator=operator, where=where, traffic=traffic or 'интернет-трафик')

    items = []
    if traffic: items.append(f'Объём трафика: {traffic}')
    if days_text: items.append(f'Срок действия: {days_text}')
    items.append(f'Оператор: {operator}')
    if country: items.append(f'Регион: {country}')
    items.append('Мгновенная доставка кода на электронную почту')
    items.append('Трафик доступен сразу после активации')
    items_html = '\n'.join(f'<li>{i}</li>' for i in items)

    notes = []
    if country:
        notes.append(f'Пакет действует только {where}')
    notes.append(f'Требуется активная SIM-карта оператора {operator}')
    if days_text:
        notes.append(f'Трафик действителен в течение {days_text} с момента активации')
    notes.append('Код активации одноразовый')
    notes_html = '\n'.join(f'<li>{n}</li>' for n in notes)

    return f"""<h2>{h2}</h2>
<p>{intro}</p>
<p><b>Параметры пакета:</b></p>
<ul>
{items_html}
</ul>
<p><b>Обратите внимание:</b></p>
<ul>
{notes_html}
</ul>

{data_instruction(operator, country)}"""


def main():
    src = os.path.join(os.path.dirname(__file__), '..', '..', 'new table', '44_filled.xlsx')
    src = os.path.abspath(src)

    print(f'Читаю: {src}')
    wb = openpyxl.load_workbook(src)
    ws = wb['Данные о товарах']

    stats = {'total': 0, 'esim': 0, 'voucher': 0, 'data': 0}

    for row_idx in range(8, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=1).value
        if not sku or not str(sku).startswith('GFT'):
            continue

        stats['total'] += 1
        idx = stats['total']

        cur_name = str(ws.cell(row=row_idx, column=7).value or '')
        brand = str(ws.cell(row=row_idx, column=11).value or '')
        old_desc = str(ws.cell(row=row_idx, column=10).value or '')

        country = extract_country(brand, old_desc)
        operator = extract_operator(brand)
        is_esim = 'esim' in brand.lower()
        is_data = 'Пакет мобильного интернета' in cur_name

        if is_esim:
            traffic = extract_traffic(cur_name)
            days_text = extract_days_from_name(cur_name)
            desc = gen_esim(country, traffic, days_text, idx)
            stats['esim'] += 1
        elif is_data:
            traffic = extract_traffic(cur_name)
            days_text = extract_days_from_name(cur_name)
            desc = gen_data(operator, country, traffic, days_text, idx)
            stats['data'] += 1
        else:
            currency, amount = extract_denomination(old_desc)
            if not currency:
                m = re.search(r'([A-Z]{3})\s*([\d.]+)', old_desc)
                if m: currency, amount = m.group(1), m.group(2)
            desc = gen_voucher(operator, country, currency, amount, idx)
            stats['voucher'] += 1

        ws.cell(row=row_idx, column=10).value = desc

    print(f'Сохраняю: {src}')
    wb.save(src)
    wb.close()

    print(f'\nВсего: {stats["total"]}')
    print(f'  eSIM: {stats["esim"]}, Ваучеры: {stats["voucher"]}, Дата-пакеты: {stats["data"]}')


if __name__ == '__main__':
    main()
