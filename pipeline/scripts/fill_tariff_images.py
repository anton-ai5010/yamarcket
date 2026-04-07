#!/usr/bin/env python3
"""
Заполняет колонку "Ссылка на изображение" (col 8) в 44_filled.xlsx.
Добавляет к каждому товару:
  - 4 универсальные карточки (email/delivery/feedback/support) с ротацией вариантов
  - 1 инструкционное изображение (eSIM или mobile_voucher)

Главное фото (main card) НЕ добавляется — останется добавить отдельно.

Порядок изображений:
  1. [ПУСТО — место для главной карточки]
  2. email_example_card_vNN.png
  3. delivery_card_vNN.png
  4. feedback_card_vNN.png
  5. support_card_vNN.png
  6. instr_esim.png или instr_mobile_voucher.png
"""

import openpyxl
import re
import os

# === URL ШАБЛОНЫ ===
YC_BASE = 'https://storage.yandexcloud.net/yamarcket'

UNIVERSAL_CARDS = [
    'email_example_card',
    'delivery_card',
    'feedback_card',
    'support_card',
]

# Инструкционные изображения в S3
INSTR_ESIM = f'{YC_BASE}/instructions_v2/instr_universal.png'
INSTR_MOBILE = f'{YC_BASE}/instructions_v2/instr_mobile_voucher.png'


def get_variant_num(product_index):
    """Вариант карточки для продукта (1-50, цикл)."""
    return (product_index % 50) + 1


def build_image_urls(product_index, is_esim):
    """Собрать список URL изображений для товара (без главной карточки)."""
    variant = get_variant_num(product_index)

    urls = []

    # 4 универсальные карточки
    for card in UNIVERSAL_CARDS:
        urls.append(f'{YC_BASE}/universal/{card}_v{variant:02d}.png')

    # Инструкционное изображение
    if is_esim:
        urls.append(INSTR_ESIM)
    else:
        urls.append(INSTR_MOBILE)

    return urls


def main():
    src = os.path.join(os.path.dirname(__file__), '..', '..', 'new table', '44_filled.xlsx')
    src = os.path.abspath(src)

    print(f'Читаю: {src}')
    wb = openpyxl.load_workbook(src)
    ws = wb['Данные о товарах']

    stats = {'total': 0, 'had_img': 0, 'no_img': 0, 'esim': 0, 'voucher': 0}
    product_index = 0

    for row_idx in range(8, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=1).value
        if not sku or not str(sku).startswith('GFT'):
            continue

        stats['total'] += 1
        product_index += 1

        brand = str(ws.cell(row=row_idx, column=11).value or '')
        is_esim = 'esim' in brand.lower()

        if is_esim:
            stats['esim'] += 1
        else:
            stats['voucher'] += 1

        # Получить текущее изображение
        current_img = ws.cell(row=row_idx, column=8).value
        current_img = str(current_img).strip() if current_img else ''

        if current_img and 'http' in current_img:
            stats['had_img'] += 1
        else:
            stats['no_img'] += 1
            current_img = ''

        # Собрать новые URL (без главной карточки)
        new_urls = build_image_urls(product_index, is_esim)

        # Если есть текущее изображение — оставляем как первое (главное)
        if current_img:
            # Убрать дубликаты если уже были universal/instruction URLs
            existing_urls = [u.strip() for u in current_img.split(',') if u.strip()]
            # Оставляем только первое изображение (главное), убираем старые universal/instruction
            main_img = existing_urls[0] if existing_urls else ''
            if main_img and 'universal/' not in main_img and 'instructions' not in main_img:
                all_urls = [main_img] + new_urls
            else:
                all_urls = new_urls
        else:
            # Без главной — только universal + instruction
            all_urls = new_urls

        # Записать через запятую
        ws.cell(row=row_idx, column=8).value = ','.join(all_urls)

    print(f'Сохраняю: {src}')
    wb.save(src)
    wb.close()

    print(f'\n=== Статистика ===')
    print(f'Всего: {stats["total"]}')
    print(f'  eSIM: {stats["esim"]}')
    print(f'  Ваучеры/пакеты: {stats["voucher"]}')
    print(f'  Было с изображением: {stats["had_img"]}')
    print(f'  Было без изображения: {stats["no_img"]}')
    print(f'  Теперь все имеют 5-6 URL (universal + instruction)')


if __name__ == '__main__':
    main()
