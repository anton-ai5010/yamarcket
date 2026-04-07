#!/usr/bin/env python3
"""
Заполняет характеристики (колонки 39-59) в 44.xlsx для категории "Тарифные планы".
Источник данных — названия, бренды и описания товаров внутри самого xlsx.
Паттерн заполнения — по образцу ПиплБот (Яндекс Маркет).

Колонки:
  39: Тип                          → "тарифный план"
  40: Тип связи                    → "сотовая связь"
  41: Назначение                   → "для телефона"
  42: Безлимитная телефонная связь  → "Нет" / "неприменимо"
  43: Безлимитный интернет          → "Нет" / "Да" / "неприменимо"
  44: Безлимитные SMS               → "Нет" / "неприменимо"
  45: Количество траффика в месяц   → "X ГБ" (если срок ~30 дней)
  46: Количество минут в месяц      → (если есть)
  47: Регион                        → (не заполняем — рос. регионы)
  48: Срок действия                 → "X дней" / "без ограничения"
  51: Тип симкарты                  → "E-SIM" / "SIM"
  52: Домашний регион               → Страна
  53: Включенный трафик             → "X ГБ"
  54: Включенные минуты             → (если есть)
  55: Включенные SMS                → (если есть)
  56: Доп. информация               → "для аккаунтов {страна}"
  57: Вид поставки                  → "электронный ключ"
  59: Регион использования          → Страна
"""

import openpyxl
import re
import copy
import os

# === МАППИНГ СТРАН ===
# Из кода бренда (| SA) или названия страны в бренде → русское название
COUNTRY_MAP = {
    # Из суффикса бренда (| XX)
    'SA': 'Саудовская Аравия',
    'AE': 'ОАЭ',
    'UK': 'Великобритания',
    'FR': 'Франция',
    'DE': 'Германия',
    'NL': 'Нидерланды',
    'OM': 'Оман',
    'KW': 'Кувейт',
    'JO': 'Иордания',
    # Из названия страны в бренде
    'Afghanistan': 'Афганистан',
    'Algeria': 'Алжир',
    'Angola': 'Ангола',
    'Austria': 'Австрия',
    'Bahrain': 'Бахрейн',
    'Bangladesh': 'Бангладеш',
    'Belgium': 'Бельгия',
    'China': 'Китай',
    'Cyprus': 'Кипр',
    'Egypt': 'Египет',
    'France': 'Франция',
    'Georgia': 'Грузия',
    'Germany': 'Германия',
    'Greece': 'Греция',
    'Guinea': 'Гвинея',
    'Hong Kong': 'Гонконг',
    'Indonesia': 'Индонезия',
    'Italy': 'Италия',
    'Japan': 'Япония',
    'Montenegro': 'Черногория',
    'Portugal': 'Португалия',
    'Qatar': 'Катар',
    'Russia': 'Россия',
    'Saudi Arabia': 'Саудовская Аравия',
    'Serbia': 'Сербия',
    'Singapore': 'Сингапур',
    'Spain': 'Испания',
    'Thailand': 'Таиланд',
    'Tunisia': 'Тунис',
    'Turkey': 'Турция',
    'United Arab Emirates': 'ОАЭ',
    'United Kingdom': 'Великобритания',
    'United States': 'США',
    'Uzbekistan': 'Узбекистан',
    'Vietnam': 'Вьетнам',
}

# Для брендов вида "Something Country" — однословные страны
SINGLE_WORD_COUNTRIES = {
    'Afghanistan': 'Афганистан',
    'Algeria': 'Алжир',
    'Angola': 'Ангола',
    'Austria': 'Австрия',
    'Bahrain': 'Бахрейн',
    'Bangladesh': 'Бангладеш',
    'Belgium': 'Бельгия',
    'China': 'Китай',
    'Cyprus': 'Кипр',
    'Egypt': 'Египет',
    'France': 'Франция',
    'Georgia': 'Грузия',
    'Germany': 'Германия',
    'Greece': 'Греция',
    'Guinea': 'Гвинея',
    'Indonesia': 'Индонезия',
    'Italy': 'Италия',
    'Japan': 'Япония',
    'Montenegro': 'Черногория',
    'Portugal': 'Португалия',
    'Qatar': 'Катар',
    'Russia': 'Россия',
    'Serbia': 'Сербия',
    'Singapore': 'Сингапур',
    'Spain': 'Испания',
    'Thailand': 'Таиланд',
    'Tunisia': 'Тунис',
    'Turkey': 'Турция',
    'Uzbekistan': 'Узбекистан',
    'Vietnam': 'Вьетнам',
}

# Двусловные страны в брендах
MULTI_WORD_COUNTRIES = {
    'Hong Kong': 'Гонконг',
    'Saudi Arabia': 'Саудовская Аравия',
    'United Arab Emirates': 'ОАЭ',
    'United Kingdom': 'Великобритания',
    'United States': 'США',
}

# Код страны из описания (Страна: XX)
DESC_COUNTRY_CODES = {
    'AF': 'Афганистан',
    'AE': 'ОАЭ',
    'DZ': 'Алжир',
    'AO': 'Ангола',
    'AT': 'Австрия',
    'BH': 'Бахрейн',
    'BD': 'Бангладеш',
    'BE': 'Бельгия',
    'CN': 'Китай',
    'CY': 'Кипр',
    'EG': 'Египет',
    'FR': 'Франция',
    'GE': 'Грузия',
    'DE': 'Германия',
    'GR': 'Греция',
    'GN': 'Гвинея',
    'HK': 'Гонконг',
    'ID': 'Индонезия',
    'IT': 'Италия',
    'JP': 'Япония',
    'JO': 'Иордания',
    'KW': 'Кувейт',
    'ME': 'Черногория',
    'OM': 'Оман',
    'PT': 'Португалия',
    'QA': 'Катар',
    'RU': 'Россия',
    'SA': 'Саудовская Аравия',
    'RS': 'Сербия',
    'SG': 'Сингапур',
    'ES': 'Испания',
    'TH': 'Таиланд',
    'TN': 'Тунис',
    'TR': 'Турция',
    'GB': 'Великобритания',
    'US': 'США',
    'UZ': 'Узбекистан',
    'VN': 'Вьетнам',
}


def extract_country(brand, name, desc):
    """Извлечь страну из бренда, названия или описания."""
    # 1. Из суффикса бренда: "Something | SA"
    m = re.search(r'\|\s*(\w+)$', brand)
    if m:
        code = m.group(1)
        if code in COUNTRY_MAP:
            return COUNTRY_MAP[code]

    # 2. Из двусловной страны в бренде
    for eng, rus in MULTI_WORD_COUNTRIES.items():
        if eng in brand:
            return rus

    # 3. Из однословной страны в бренде (последнее слово)
    brand_words = brand.split()
    if len(brand_words) >= 2:
        last_word = brand_words[-1]
        if last_word in SINGLE_WORD_COUNTRIES:
            return SINGLE_WORD_COUNTRIES[last_word]

    # 4. Из кода страны в описании (Страна: XX)
    m = re.search(r'Страна:\s*(\w+)', desc)
    if m:
        code = m.group(1)
        if code in DESC_COUNTRY_CODES:
            return DESC_COUNTRY_CODES[code]

    # 5. Из названия — ищем русские названия стран
    ru_countries = [
        'Саудовской Аравии', 'Саудовская Аравия',
        'Великобритании', 'Великобритания',
        'ОАЭ', 'США', 'Франции', 'Франция',
        'Германии', 'Германия', 'Турции', 'Турция',
    ]
    for c in ru_countries:
        if c in name:
            # Normalize to nominative
            norm = {
                'Саудовской Аравии': 'Саудовская Аравия',
                'Великобритании': 'Великобритания',
                'Франции': 'Франция',
                'Германии': 'Германия',
                'Турции': 'Турция',
            }
            return norm.get(c, c)

    # 6. Airalo — международный роуминг
    if 'airalo' in brand.lower():
        return 'весь мир'

    return None


def is_esim(brand, name):
    """Определить, является ли товар eSIM."""
    combined = (brand + ' ' + name).lower()
    return 'esim' in combined or 'e-sim' in combined


def extract_traffic_gb(name, desc):
    """Извлечь объём трафика в ГБ из названия или описания."""
    combined = name + ' ' + desc
    # Patterns: "10GB", "1.5GB", "2ГБ", "1 GB", "75GB", "Unlimited"
    m = re.search(r'([\d.]+)\s*(?:GB|ГБ)', combined, re.IGNORECASE)
    if m:
        return m.group(1)  # return as string to preserve decimals

    # "Unlimited" — безлимит
    if re.search(r'unlimited|безлимит', combined, re.IGNORECASE):
        return 'unlimited'

    return None


def extract_days(name, desc):
    """Извлечь срок действия в днях из названия или описания."""
    combined = name + ' ' + desc
    # "7 Days", "30 Days", "15 дней"
    m = re.search(r'(\d+)\s*(?:Days?|дн)', combined, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def extract_minutes(name, desc):
    """Извлечь количество минут."""
    combined = name + ' ' + desc
    m = re.search(r'(\d+)\s*(?:мин|min)', combined, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def format_traffic(gb):
    """Форматировать трафик для колонки."""
    if gb is None:
        return None
    if gb == -1:
        return None  # безлимит указывается отдельно
    return f'{gb} ГБ'


def format_days(days):
    """Форматировать срок действия."""
    if days is None:
        return None
    if days == 1:
        return '1 день'
    elif days in (2, 3, 4):
        return f'{days} дня'
    else:
        return f'{days} дней'


def main():
    src = os.path.join(os.path.dirname(__file__), '..', '..', 'new table', '44.xlsx')
    src = os.path.abspath(src)
    dst = src.replace('44.xlsx', '44_filled.xlsx')

    print(f'Читаю: {src}')
    wb = openpyxl.load_workbook(src)
    ws = wb['Данные о товарах']

    # Column indices (0-based in our data, but openpyxl is 1-based)
    # Col 39 = Тип → column AN (40th, 1-based)
    COL = {
        'type': 40,            # 39 (0-based) → col 40 (1-based) = Тип
        'conn_type': 41,       # Тип связи
        'purpose': 42,         # Назначение
        'unlim_phone': 43,     # Безлимитная телефонная связь
        'unlim_internet': 44,  # Безлимитный интернет
        'unlim_sms': 45,       # Безлимитные SMS
        'traffic_month': 46,   # Количество траффика в месяц
        'minutes_month': 47,   # Количество минут в месяц
        'region': 48,          # Регион
        'validity': 49,        # Срок действия
        'number_type': 50,     # Тип номера
        'cell_standard': 51,   # Стандарт сотовой связи
        'sim_type': 52,        # Тип симкарты
        'home_region': 53,     # Домашний регион
        'incl_traffic': 54,    # Количество включенного траффика
        'incl_minutes': 55,    # Количество включенных минут
        'incl_sms': 56,        # Количество SMS/MMS
        'extra_info': 57,      # Дополнительная информация
        'delivery': 58,        # Вид поставки
        'pkg_count': 59,       # Количество упаковок
        'use_region': 60,      # Регион использования
    }

    stats = {'total': 0, 'esim': 0, 'voucher': 0, 'country_found': 0, 'traffic_found': 0, 'days_found': 0}

    for row_idx in range(8, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=1).value
        if not sku or not str(sku).startswith('GFT'):
            continue

        name = str(ws.cell(row=row_idx, column=7).value or '')
        brand = str(ws.cell(row=row_idx, column=11).value or '')
        desc = str(ws.cell(row=row_idx, column=10).value or '')

        stats['total'] += 1

        # Определяем тип
        esim = is_esim(brand, name)
        if esim:
            stats['esim'] += 1
        else:
            stats['voucher'] += 1

        # Извлекаем данные
        country = extract_country(brand, name, desc)
        traffic_gb = extract_traffic_gb(name, desc)
        days = extract_days(name, desc)
        minutes = extract_minutes(name, desc)
        is_unlimited = traffic_gb == 'unlimited'

        if country:
            stats['country_found'] += 1
        if traffic_gb is not None and traffic_gb != 'unlimited':
            stats['traffic_found'] += 1
        if days is not None:
            stats['days_found'] += 1

        # === ЗАПОЛНЯЕМ ХАРАКТЕРИСТИКИ ===

        # Тип → всегда "тарифный план"
        ws.cell(row=row_idx, column=COL['type']).value = 'тарифный план'

        # Тип связи → "сотовая связь"
        ws.cell(row=row_idx, column=COL['conn_type']).value = 'сотовая связь'

        # Назначение → "для телефона"
        ws.cell(row=row_idx, column=COL['purpose']).value = 'для телефона'

        # Безлимитная телефонная связь
        if esim:
            # eSIM data-only → Нет
            ws.cell(row=row_idx, column=COL['unlim_phone']).value = 'Нет'
        else:
            # Ваучеры — неприменимо (мы не знаем)
            ws.cell(row=row_idx, column=COL['unlim_phone']).value = 'неприменимо'

        # Безлимитный интернет
        if is_unlimited:
            ws.cell(row=row_idx, column=COL['unlim_internet']).value = 'Да'
        elif esim:
            ws.cell(row=row_idx, column=COL['unlim_internet']).value = 'Нет'
        else:
            ws.cell(row=row_idx, column=COL['unlim_internet']).value = 'неприменимо'

        # Безлимитные SMS
        if esim:
            ws.cell(row=row_idx, column=COL['unlim_sms']).value = 'Нет'
        else:
            ws.cell(row=row_idx, column=COL['unlim_sms']).value = 'неприменимо'

        # Трафик в месяц (только если срок ~30 дней)
        if traffic_gb and traffic_gb != 'unlimited' and days and 28 <= days <= 31:
            ws.cell(row=row_idx, column=COL['traffic_month']).value = f'{traffic_gb} ГБ'

        # Минуты в месяц
        if minutes and days and 28 <= days <= 31:
            ws.cell(row=row_idx, column=COL['minutes_month']).value = minutes

        # Срок действия
        if days:
            ws.cell(row=row_idx, column=COL['validity']).value = format_days(days)

        # Тип симкарты
        if esim:
            ws.cell(row=row_idx, column=COL['sim_type']).value = 'E-SIM'
        # Для ваучеров не заполняем — это не SIM

        # Домашний регион
        if country:
            ws.cell(row=row_idx, column=COL['home_region']).value = country

        # Включенный трафик (строка "X ГБ" — как у ПиплБот)
        if traffic_gb and traffic_gb != 'unlimited':
            ws.cell(row=row_idx, column=COL['incl_traffic']).value = f'{traffic_gb} ГБ'

        # Включенные минуты
        if minutes:
            ws.cell(row=row_idx, column=COL['incl_minutes']).value = minutes

        # Дополнительная информация
        if country:
            ws.cell(row=row_idx, column=COL['extra_info']).value = f'для аккаунтов {country}'

        # Вид поставки → электронный ключ
        ws.cell(row=row_idx, column=COL['delivery']).value = 'электронный ключ'

        # Регион использования
        if country:
            ws.cell(row=row_idx, column=COL['use_region']).value = country

    # Сохраняем
    print(f'\nСохраняю: {dst}')
    wb.save(dst)
    wb.close()

    print(f'\n=== Статистика ===')
    print(f'Всего обработано: {stats["total"]}')
    print(f'  eSIM: {stats["esim"]}')
    print(f'  Ваучеры: {stats["voucher"]}')
    print(f'  Страна найдена: {stats["country_found"]}/{stats["total"]}')
    print(f'  Трафик найден: {stats["traffic_found"]}/{stats["total"]}')
    print(f'  Срок найден: {stats["days_found"]}/{stats["total"]}')


if __name__ == '__main__':
    main()
