"""
sync_json_to_xlsx.py

Переносит все правки из JSON (источник истины) в 2.xlsx (файл Яндекс Маркета).

Обновляет: название, изображения, описание, бренд, теги, все характеристики.
НЕ трогает: цены, ошибки Маркета, качество карточки, статус архива.
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent.parent
XLSX_PATH = ROOT / "new table" / "2.xlsx"
JSON_PATH = ROOT / "new table" / "Онлайн-подписки и карты оплаты_истина_с_характеристиками.json"
OUT_PATH = ROOT / "new table" / "2_updated.xlsx"

# Колонки, которые НЕ трогаем (readonly от Маркета или управляются отдельно)
# Все имена — нормализованные (без *, без лишних пробелов)
SKIP_HEADERS = {
    "Критичные ошибки",
    "Некритичные ошибки",
    "Качество карточки",
    "Рекомендации по заполнению",
    # Цены не трогаем — управляются в ЛК
    "Цена",
    "Зачёркнутая цена",
    "Себестоимость",
    "Дополнительные расходы",
    # Архив — статус управляется вручную
    "В архиве",
    # Прочие характеристики — сложный составной формат
    "Прочие характеристики",
    "PARAM_NAMES",
    # Поля Маркета только для чтения
    "Артикул товара (SKU)",
    "CSKU на Маркете",
    "Дата дополнения карточки",
}


def load_json_data(path: Path) -> dict[str, list]:
    """Возвращает {SKU: row_list} для всех GFT-товаров."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    rows = data["Данные о товарах"]

    # Заголовки — строка 3 (индекс 3)
    headers = rows[3]

    # Данные — с 7-й строки (там где начинается GFT)
    result = {}
    for row in rows[7:]:
        sku = str(row[0] or "").strip()
        if sku.startswith("GFT"):
            result[sku] = row

    return headers, result


def main():
    print(f"Читаю JSON: {JSON_PATH}")
    json_headers, json_data = load_json_data(JSON_PATH)
    print(f"  Товаров в JSON: {len(json_data)}")

    # Строим индекс: имя_заголовка → индекс_в_json
    json_header_idx = {}
    for i, h in enumerate(json_headers):
        if h and str(h).strip():
            # Нормализуем — убираем * и лишние пробелы
            name = str(h).strip().rstrip(" *")
            json_header_idx[name] = i

    print(f"\nЧитаю XLSX: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Данные о товарах"]
    print(f"  Строк в листе: {ws.max_row}")

    # Читаем заголовки xlsx из строки 4
    xlsx_headers = {}  # col_idx (1-based) → имя
    header_row = 4
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v:
            name = str(v).strip().rstrip(" *")
            xlsx_headers[col] = name

    # Строим маппинг: xlsx_col → json_col_idx
    # Матчинг по имени заголовка
    col_mapping = {}  # xlsx_col → json_idx
    for xlsx_col, xname in xlsx_headers.items():
        if xname in SKIP_HEADERS:
            continue
        # Ищем в JSON по имени
        if xname in json_header_idx:
            col_mapping[xlsx_col] = json_header_idx[xname]

    print(f"\nСовпадений колонок: {len(col_mapping)}")

    # Данные начинаются со строки 8 в xlsx
    DATA_START_ROW = 8
    updated_rows = 0
    updated_cells = 0
    not_found = 0

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        sku_cell = ws.cell(row=row_num, column=1)
        sku = str(sku_cell.value or "").strip()

        if not sku.startswith("GFT"):
            continue

        if sku not in json_data:
            not_found += 1
            continue

        json_row = json_data[sku]
        row_changed = False

        for xlsx_col, json_idx in col_mapping.items():
            if json_idx >= len(json_row):
                continue

            json_val = json_row[json_idx]
            xlsx_cell = ws.cell(row=row_num, column=xlsx_col)
            xlsx_val = xlsx_cell.value

            # Нормализуем для сравнения
            jv = json_val if json_val is not None else None
            xv = xlsx_val if xlsx_val is not None else None

            # Конвертируем типы для сравнения
            if isinstance(jv, float) and jv == int(jv):
                jv_str = str(int(jv))
            else:
                jv_str = str(jv) if jv is not None else ""

            xv_str = str(xv) if xv is not None else ""

            if jv_str != xv_str:
                xlsx_cell.value = json_val
                updated_cells += 1
                row_changed = True

        if row_changed:
            updated_rows += 1

        if updated_rows % 500 == 0 and updated_rows > 0:
            print(f"  Обработано {updated_rows} строк...")

    print(f"\nРезультаты:")
    print(f"  Обновлено строк:  {updated_rows}")
    print(f"  Обновлено ячеек: {updated_cells}")
    print(f"  Не найдено в JSON: {not_found}")

    print(f"\nСохраняю в {OUT_PATH}...")
    wb.save(OUT_PATH)
    print("Готово!")


if __name__ == "__main__":
    main()
