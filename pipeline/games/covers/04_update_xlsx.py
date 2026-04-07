"""
04_update_xlsx.py
Добавляет URL обложек игр/DLC в 1.xlsx (колонка H = Ссылка на изображение).

Логика:
- Для каждого игрового SKU (приложение=сервис активации, назначение=игры)
  вставляет обложку как ПЕРВОЕ изображение
- Остальные изображения (инструкции, etc.) сдвигаются на позиции 2+
- Если обложка уже есть (URL содержит games/) — пропускаем

Формат колонки H в 1.xlsx: URL через запятую (до 30 штук)
"""
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent.parent.parent
XLSX_PATH = ROOT / "new table" / "1.xlsx"
MAPPING_PATH = ROOT / "pipeline" / "images" / "output" / "games_covers" / "sku_mapping.json"

BASE_URL = "https://raw.githubusercontent.com/anton-ai5010/yamarcket-images/main/games/"

COL_SKU = 1       # A
COL_IMAGES = 8    # H — Ссылка на изображение
DATA_START_ROW = 8


def get_cover_url(slug: str) -> str:
    return BASE_URL + slug + ".jpg"


def main():
    print(f"Читаю {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Находим лист с товарами
    sheet_name = None
    for name in wb.sheetnames:
        if "товар" in name.lower() or "данные" in name.lower() or name == "Список товаров":
            sheet_name = name
            break
    if not sheet_name:
        # Берём первый рабочий лист (не Инструкция и не Enums)
        for name in wb.sheetnames:
            if name not in ("Инструкция", "Enums", "Описание полей", "Настройки",
                            "Таблицы размеров", "Список значений", "Размерные сетки",
                            "Информация по размерным сеткам", "MeasureUnitsMeta", "Mapping"):
                sheet_name = name
                break

    ws = wb[sheet_name]
    print(f"Лист: {sheet_name!r}")

    with open(MAPPING_PATH) as f:
        mapping = json.load(f)

    updated = 0
    skipped_no_cover = 0
    skipped_already = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        sku_cell = ws.cell(row=row_idx, column=COL_SKU)
        sku = str(sku_cell.value or "").strip()

        if not sku or not sku.startswith("GFT"):
            continue

        if sku not in mapping:
            continue

        cover_info = mapping[sku]
        if not cover_info.get("exists"):
            skipped_no_cover += 1
            continue

        cover_url = get_cover_url(cover_info["slug"])

        # Читаем текущие изображения
        images_cell = ws.cell(row=row_idx, column=COL_IMAGES)
        current = str(images_cell.value or "").strip()

        # Если обложка уже есть — пропускаем
        if cover_url in current or "/games/" in current:
            skipped_already += 1
            continue

        # Добавляем обложку как первое фото
        if current:
            new_value = cover_url + "," + current
        else:
            new_value = cover_url

        # Ограничиваем 30 ссылками
        urls = [u.strip() for u in new_value.split(",") if u.strip()][:30]
        images_cell.value = ",".join(urls)
        updated += 1

    print(f"Обновлено:    {updated}")
    print(f"Уже есть:     {skipped_already}")
    print(f"Нет обложки:  {skipped_no_cover}")

    wb.save(XLSX_PATH)
    print(f"Сохранено:    {XLSX_PATH}")


if __name__ == "__main__":
    main()
